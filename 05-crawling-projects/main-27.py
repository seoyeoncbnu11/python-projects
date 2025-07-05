import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re
from tqdm import tqdm


filePath = "./05-crawling-projects/고등교육기관 하반기 주소록(2023).xlsx"
df_from_excel = pd.read_excel(filePath, engine="openpyxl")
df_from_excel.columns = df_from_excel.loc[4].tolist()
df_from_excel = df_from_excel.drop(index=list(range(0, 5)))

# print(df_from_excel.head())
# print(df_from_excel["학교명"].values)
# print(df_from_excel["주소"].values)

import requests
import requests

apiurl = "https://api.vworld.kr/req/address?"
params = {
    "service": "address",
    "request": "getcoord",
    "crs": "epsg:4326",
    "format": "json",
    "type": "road",
    "key": "55E673C6-6251-30B4-8249-284999BD6C87",
}


def request_geo(road):
    params["address"] = road
    page = requests.get(apiurl, params=params)
    json_data = page.json()
    if json_data["response"]["status"] == "OK":
        x = json_data["response"]["result"]["point"]["x"]
        y = json_data["response"]["result"]["point"]["y"]
        return x, y
    else:
        x = 0
        y = 0
        return x, y


try:
    wb = load_workbook("./05-crawling-projects/학교주소좌표.xlsx", data_only=True)
    sheet = wb.active

except:
    wb = Workbook()
    sheet = wb.active

university_list = df_from_excel["학교명"].to_list()
address_list = df_from_excel["주소"].to_list()

for num, value in tqdm(enumerate(address_list)):
    addr = re.sub(r"\([^)]*\)", "", value)
    x, y = request_geo(addr)
    sheet.append([university_list[num], addr, x, y])

wb.save("./05-crawling-projects/학교주소좌표.xlsx")
